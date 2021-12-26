
# written by seyed mohamad mahdi moravej

# remember to use ctrl-shift-number and
# ctrl-shift-k for number bookmark extension

# -------------------------------import necessary module-----------
import jdatetime
from openpyxl import *
from telegram import *
from telegram.ext import *
from random import *
from jdatetime import *
# -------------------------------define overal variable------------
token = "2076371567:AAF7xXQMPI86FpcFWZVgWzlK6jtWKb60DDw"
bot = Bot(token)
wb = Workbook()
# -------------------------------define overal list-----------------
# list_name_month = ["farvardin","ordibehesht","khordad","tir","mordad","shahrivar","mehr","aban","azar","dey","bahman","esfand"]
list_name_bakhshha = ["jafarie","dastjerd","salafchegan","kahak","markazi"]
list_mechanical = ["36020104.xlsx", "36020111.xlsx", "36020115.xlsx",
                    "041.xlsx", "042.xlsx", "047.xlsx", "048.xlsx", 
                    "049.xlsx", "051.xlsx", "053.xlsx", 
                    "36020503.xlsx", "36020504.xlsx", "36020505.xlsx",
                    "36020506.xlsx", "36020507.xlsx"
                    ]
dict_month={"Farvardin":"فروردین","Ordibehesht":"اردیبهشت","Khordad":"خرداد",
           "Tir":"تیر","Mordad":"مرداد","Shahrivar":"شهریور",
           "Mehr":"مهر","Aban":"آبان","Azar":"آذر",
           "Dey":"دی","Bahman":"بهمن","Esfand":"اسفند"}
list_electrical = ["043.xlsx", "044.xlsx", "045.xlsx",
                   "046.xlsx", "050.xlsx", "052.xlsx"]
list_administrator = [
                    "36010101.xlsx",
                    "36010102.xlsx", "36010103.xlsx",
                    "36010201.xlsx", "36010501.xlsx", "36010601.xlsx",
                    "36010602.xlsx", "36010801.xlsx", "36010802.xlsx",
                    "36020101.xlsx", "36020109.xlsx", "36020112.xlsx",
                    "36020114.xlsx", "36020116.xlsx", "36020602.xlsx",
                    "36020603.xlsx", "36020604.xlsx", "36020607.xlsx",
                    "36020901.xlsx", "36020903.xlsx", "36020904.xlsx",
                    "36020905.xlsx", "36020906.xlsx", "36020907.xlsx",
                    "36020909.xlsx", "36020911.xlsx", "360201106.xlsx",
                    "360201107.xlsx", "360201108.xlsx"
                        ]
list_quality_control = ["36010704.xlsx", "36010705.xlsx", 
                        "36010706.xlsx" ''' file isn't existed but should be exist'''
                        "36020401.xlsx", "36020402.xlsx", 
                        "36020403.xlsx", "36020404.xlsx", "36020406.xlsx", "36020407.xlsx", 
                        "36020408.xlsx", "36020410.xlsx", 
                        "36020601.xlsx",
                        "36010708.xlsx", ''' file isn't existed'''
                        "36020703.xlsx", ''' file isn't existed'''
                        "36020801.xlsx", 
                        "36020802.xlsx", ''' file isn't existed'''
                        "36032004.xlsx", ''' file isn't existed'''
                        ]
list_water_without_income = ["36021001.xlsx", "360201105.xlsx"]
name_bakhshha = ["jafarie", "dastjerd", "salafchegan", "kahak", "markazi"]
# -------------------------------define class-----------------------
class Mybot():
# -----------------------------------define __init__----------------
    def __init__(self):
# -----------------------------------------define keyboard ---------
# ---------------------------------------------define your contractor(first_query)
        keyboard_choose_paymankar = [
                    [InlineKeyboardButton(
                        "آقای فتحی", 
                        callback_data="فتحی")], 
                    [InlineKeyboardButton(
                        "آقای مهدوی",
                        callback_data="mahdavi")]
                    ]
# ---------------------------------------------define your time to do-------
        keyboard_choose_time = [
                    [InlineKeyboardButton(
                    "فروردین", 
                    callback_data="Farvardin")],
                    [InlineKeyboardButton(
                    "اردیبهشت", 
                    callback_data="Ordibehesht")],
                    [InlineKeyboardButton(
                    "خرداد", 
                    callback_data="Khordad")],
                    [InlineKeyboardButton(
                    "تیر", 
                    callback_data="Tir")],
                    [InlineKeyboardButton(
                    "مرداد", 
                    callback_data="Mordad")],
                    [InlineKeyboardButton(
                    "شهریور", 
                    callback_data="Shahrivar")],
                    [InlineKeyboardButton(
                    "مهر", 
                    callback_data="Mehr")],
                    [InlineKeyboardButton(
                    "آبان", 
                    callback_data="Aban")],
                    [InlineKeyboardButton(
                    "آذر", 
                    callback_data="Azar")],
                    [InlineKeyboardButton(
                    "دی", 
                    callback_data="Dey")],
                    [InlineKeyboardButton(
                    "بهمن", 
                    callback_data="Bahman")],
                    [InlineKeyboardButton(
                    "اسفند", 
                    callback_data="Esfand")],
                    [InlineKeyboardButton(
                    "خام", 
                    callback_data="kham")]                    
                    ]
# --------------------------------  chossing location to fill_online
        keyboard_name_bakhshha = [
                        [InlineKeyboardButton(
                        "جعفریه", 
                        callback_data="jafarie")],
                        [InlineKeyboardButton(
                        "دستجرد", 
                        callback_data="dastjerd")],
                        [InlineKeyboardButton(
                        "سلفچگان", 
                        callback_data="salafchegan")],
                        [InlineKeyboardButton(
                        "کهک", 
                        callback_data="kahak")],
                        [InlineKeyboardButton(
                        "مرکزی", 
                        callback_data="markazi")]
                        ]
        keyboard_choose_office = [
                    [InlineKeyboardButton(
                    "واحد فنی مکانیکی", 
                    callback_data="fill_online_mechanical")], 
                    [InlineKeyboardButton(
                    "واحد فنی برقی", 
                    callback_data="fill_online_electrical")], 
                    [InlineKeyboardButton(
                    "آبدار ،ضایعات گیر و سایر نیروهای مرتبط با سرپرست بخش", 
                    callback_data="fill_online_administrator")], 
                    [InlineKeyboardButton(
                    "کنترل کیفی", 
                    callback_data="fill_online_quality_control")], 
                    [InlineKeyboardButton(
                    "واحد دبی سنجی و آب بدون درآمد ", 
                    callback_data="fill_online_debi_gage")]
                        ]
# chossing relevant mechanical file to fill_online
        keyboard_fill_online_mechanical = [
                    [InlineKeyboardButton(
                    "فرم بازدید و مانور شیرآلات کنترل دبی و فشار", 
                    callback_data="36020104")], 
                    [InlineKeyboardButton(
                    "فرم بازدید و مانور شیرآلات فلوتری", 
                    callback_data="36020111")], 
                    [InlineKeyboardButton(
                    "فرم نگهداشت برنامه‌ای و سرویس عملگرهای برقی", 
                    callback_data="36020115")], 
                    [InlineKeyboardButton(
                    "فرم نگهداشت برنامه‌ای و بررسی عملکرد الکتروپمپ شناور", 
                    callback_data="36020503")], 
                    [InlineKeyboardButton(
                    "فرم نگهداشت برنامه‌ای و بررسی عملکرد جرثقیل سقفی", 
                    callback_data="36020504")], 
                    [InlineKeyboardButton(
                    "فرم نگهداشت برنامه‌ای و بررسی عملکرد کمپرسور هوا", 
                    callback_data="36020505")], 
                    [InlineKeyboardButton(
                    "فرم نگهداشت برنامه‌ای و سرویس عملکرد الکتروپمپ زمینی", 
                    callback_data="36020506")], 
                    [InlineKeyboardButton(
                    "فرم ترموگرافی و ارتعاش‌سنجی الکتروپمپ", 
                    callback_data="36020507")], 
                    [InlineKeyboardButton(
                    "چک لیست تابلوهای برق", 
                    callback_data="041")], 
                    [InlineKeyboardButton(
                    "چک لیست تجهیزات ایستگاه پمپاژ", 
                    callback_data="042")], 
                    [InlineKeyboardButton(
                    "فرم فعاليتهاي پيشگيرانه دیزل ژنراتور (در حالت سرد)", 
                    callback_data="047")], 
                    [InlineKeyboardButton(
                    "جدول کارکرد دیزل ژنراتورها", 
                    callback_data="048")], 
                    [InlineKeyboardButton(
                    "دستور العمل فعاليتهاي پيشگيرانه (در حالت سرد)", 
                    callback_data="049")], 
                    [InlineKeyboardButton(
                    "چک لیست نظارتی عملیات ایرلیف (لایروبی) و احیاء", 
                    callback_data="051")], 
                    [InlineKeyboardButton(
                    "چک لیست محاسبه راندمان", 
                    callback_data="053")]
                    ]
# choosing relevant electrical file to fill_online
        keyboard_fill_online_electrical = [
                    [InlineKeyboardButton(
                    "گزارش ماهانه قطعات مصرفی تاسیسات برق", 
                    callback_data="043")], 
                    [InlineKeyboardButton(
                    "فرم فعالیت نصب و راه اندازی الکتروپمپ", 
                    callback_data="044")], 
                    [InlineKeyboardButton(
                    "فرم گزارش روزانه تاسیسات برقی", 
                    callback_data="045")], 
                    [InlineKeyboardButton(
                    "چک لیست مواد مصرفی الکتریکی", 
                    callback_data="046")], 
                    [InlineKeyboardButton(
                    "چک لیست جمع آوری تجهیزات چاه", 
                    callback_data="050")], 
                    [InlineKeyboardButton(
                    "نصب و راه اندازی الکتروپمپ شناور", 
                    callback_data="052")]
                    ]
# choosing relevant administrator file
        keyboard_fill_online_administrator = [
                    [InlineKeyboardButton(
                    "فرم راهبری و اپراتوری چاه", 
                    callback_data="36010101.xlsx")], 
                    [InlineKeyboardButton(
                    "فرم راهبری و اپراتوری چشمه", 
                    callback_data="36010102")
                    ], 
                    [InlineKeyboardButton(
                    "فرم راهبری و اپراتوری قنات", 
                    callback_data="36010103")
                    ], 
                    [InlineKeyboardButton(
                    "فرم راهبری و اپراتوری تاسیسات و شیرآلات ایستگاه پمپاژ ", 
                    callback_data="36010201")
                    ], 
                    [InlineKeyboardButton(
                    "فرم راهبری و اپراتوری خطوط انتقال آب / خط بان", 
                    callback_data="36010501")
                    ], 
                    [InlineKeyboardButton(
                    "فرم راهبری و ثبت گزارش اپراتوری مخازن ذخیره آب و تاسیسات جانبی", 
                    callback_data="36010601")
                    ], 
                    [InlineKeyboardButton(
                    "فرم راهبری و اپراتوری شبکه توزیع آب", 
                    callback_data="36010602")
                    ], 
                    [InlineKeyboardButton(
                    "فرم انجام خدمات فضای سبز", 
                    callback_data="36010801")
                    ], 
                    # [InlineKeyboardButton(
                    # #  negahban fiziki por kone
                    # "حفاظت از تاسیسات آب شرب", 
                    # callback_data="36010802")
                    # ], 
                    # [InlineKeyboardButton(
                    # #   shir haye shabake tozi va khate enteqal
                    # "فرم بازدید و مانور شیرآلات", 
                    # callback_data="36020101")
                    # ], 
                    # [InlineKeyboardButton(
                    # # dasture kari ba bakhsh ha karfarma mide
                    # "فرم پیدا کردن شیرآلات ناپیدا به وسیله حفاری", 
                    # callback_data="36020109")
                    # ], 
                    # [InlineKeyboardButton(
                    # # safi khate_enteql_va _shabake tozi
                    # "فرم بازدید و سرویس صافی شیرآلات", 
                    # callback_data="36020112")
                    # ], 
                    # [InlineKeyboardButton(
                    # # shabake tozi va khate enteqal
                    # "فرم بازدید و مانور شیرآلات هوا", 
                    # callback_data="36020114")
                    # ], 
                    # [InlineKeyboardButton(
                    # #   moredi
                    # "فرم باز و بسته نمودن شیرآلات به منظور نوبت بندی", 
                    # callback_data="36020116")
                    # ], 
                    # [InlineKeyboardButton(
                    # # dasture kari ba bakhsh ha karfarma mide
                    # "فرم نظافت و نگهداشت برنامه‌ای محوطه تاسیسات آب محصور شده", 
                    # callback_data="36020602")
                    # ],    
                    # [InlineKeyboardButton(
                    # # dasture kari ba bakhsh ha karfarma mide
                    # "فرم نظافت و نگهداشت حریم منابع آبی و مخازن غیرمحصور", 
                    # callback_data="36020603")
                    # ],      
                    # [InlineKeyboardButton(
                    # #rutin
                    # "فرم نگهداشت برنامه‌ای ونظافت مستمر ساختمان‌های تاسیسات مکانیکی و برقی", 
                    # callback_data="36020604")
                    # ],                                                                                                                                                          
                    # [InlineKeyboardButton(
                    # # hozchehaye shabake tozi va khate enteqal
                    # "فرم نگهداشت برنامه‌ای و نظافت حوضچه", 
                    # callback_data="36020607")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم شستشو و گندزدایی مخازن زمینی", 
                    # callback_data="36020901")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم شستشو و گندزدایی خطوط انتقال", 
                    # callback_data="36020903")
                    # ], 
                    # [InlineKeyboardButton(
                    # #nadarim
                    # "فرم شستشو و گندزدایی خطوط انتقال با آب و هوا", 
                    # callback_data="36020904")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم شستشو و گندزدایی شبکه توزیع", 
                    # callback_data="36020905")
                    # ],                                                                 
                    # [InlineKeyboardButton(
                    # "فرم شستشو و گندزدایی مخازن هوایی", 
                    # callback_data="36020906")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم لایروبی مخازن ذخیره آب و حمل لای با هر نوع وسیله مکانیکی", 
                    # callback_data="36020907")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم شستشو تانکرهای آب رسانی", 
                    # callback_data="36020909")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم رسوب‌زدایی خطوط انتقال بدون دستگاه رسوب‌زدایی براساس دستورالعمل", 
                    # callback_data="36020911")
                    # ], 
                    # [InlineKeyboardButton(
                    # # dasture kari ba bakhsh ha karfarma mide
                    # "فرم مرئی‌سازی شیرآلات حوضچه", 
                    # callback_data="360201106")
                    # ],     
                    # [InlineKeyboardButton(
                    # # dasture kari ba bakhsh ha karfarma mide
                    # "فرم هم‌سطح سازی حوضچه شیرآلات", 
                    # callback_data="360201107")
                    # ], 
                    # [InlineKeyboardButton(
                    # # dasture kari ba bakhsh ha karfarma mide
                    # "فرم پیدا کردن شیرآلات ناپیدا", 
                    # callback_data="360201108")
                    # ]
                    ]
# choosing relevant measuring quality control
        keyboard_fill_online_quality_control = [
                    # [InlineKeyboardButton(
                    # "کلرسنجی", 
                    # callback_data="36010704")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم راهبری دستگاه کلریناتور گازی همراه با کلرسنجی", 
                    # callback_data="36010705")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم راهبری و اپراتوری گندزدایی دستگاه آب ژاول به همراه کلرسنجی", 
                    # callback_data="36010706")
                    # ], 
                    # [InlineKeyboardButton(
                    # #    moredi
                    # "فرم نگهداشت برنامه‌ای و سرویس دستگاه کلریناتور محلولی برقی به طور مستمر", 
                    # callback_data="36020401")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم نگهداشت برنامه‌ای و سرویس دستگاه کلریناتور محلولی به طور مستمر", 
                    # callback_data="36020402")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم نگهداشت برنامه‌ای و سرویس دستگاه کلریناتور گازی به طور مستمر", 
                    # callback_data="36020403")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم نگهداشت برنامه‌ای و سرویس الکترولیز نمک طعام به طور مستمر", 
                    # callback_data="36020404")
                    # ], 
                    # [InlineKeyboardButton(
                    # "فرم نگهداشت برنامه‌ای و سرویس سامانه ازن‌زنی به طور مستمر", 
                    # callback_data="36020406")
                    # ],  
                    # [InlineKeyboardButton(
                    # "فرم نگهداشت برنامه‌ای و سرویس سامانه خنثی‌کننده گاز کلر اسکرابر", 
                    # callback_data="36020407")
                    # ], 
                    [InlineKeyboardButton(
                    "فرم نگهداشت برنامه‌ای و سرویس سیستم تزریق آب ژاول", 
                    # mohem
                    callback_data="36020408")
                    ], 
                    # [InlineKeyboardButton(
                    # "فرم نگهداشت برنامه‌ای و سرویس سیستم تزریق مواد شیمیایی", 
                    # callback_data="36020410")
                    # ], 
                    [InlineKeyboardButton(
                    "فرم نظافت اتاق‌های کلریناتور و انبار کلر", 
                    # mohem       
                    callback_data="36020601")
                    ], 
                    # [InlineKeyboardButton(
                    # "فرم نمونه‌برداری (میکروبیولوژی و باکتریولوژی) از شبکه توزیع آب، منابع و مخازن و تصفیه‌خانه‌ها", 
                    # callback_data="36020801")
                    # ]
                    ]
# choosing relevant measuring debi and water without income file
        keyboard_fill_online_debi_gage = [
                     [InlineKeyboardButton(
                    "فرم دبی‌سنجی با دستگاه دبی‌سنج قابل حمل", 
                    callback_data="36021001")], 
                    [InlineKeyboardButton(
                    "فرم قرائت و ثبت داده دستگاه اندازه‌گیری جریان", 
                    callback_data="360201105")
                    ]
                    ]   
        keyboard_item_well = [[
                    InlineKeyboardButton(
                    "1-وضعیت صحت عملکرد الکتروپمپ", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "2-وضعیت صحت عملکرد تابلو برق", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "3-وضعیت صحت عملکرد خط رانش و تخلیه", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "4-وضعیت صحت عملکرد کابل سر چاه و سیم نگهدارنده", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "5-طریقه ی خاموش و روشن کردن الکتروپمپ ها و دستگاه های جانبی", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "6-تحویل گیری تاسیسات چاه پس از راه اندازی", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "7-قرائت مقادیرآبدهی چاه و فشار سرچاه", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "8-گزارش مقادیر ولتاژ و آمپر متوسط تابلو و cos Ф", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "9-کنترل آلودگی میکروبی در زمان سیلاب و کنترل کدورت خروجی", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "10-بررسی وضعیت نشت روغن", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "11-بررسی وضعیت صدا و لرزش غیرعادی پمپ ،شیرآلات و اتصالات", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "12-بررسی وضعیت فیوزهای کات اوت ترانس", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "13-بررسی وضعیت نظافت تابلو", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "14-بررسی وضعیت بوی سوختگی غیرعادی", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "15-بررسی وضعیت چراغ های سیگنال", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "16-بررسی وضعیت اتصالات", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "17-بررسی وضعیت دمای محیط و تهویه", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "18-بررسی وضعیت ظاهر تابلو", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "19-بررسی وضعیت اندازه گیری و ثبت آمپر و ولتاژ", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "20-بررسی ظاهری سیستم های روشنایی", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "21-وضعیت انشعابات برق", 
                    callback_data="electropump")],                      
                    [InlineKeyboardButton(
                    "22-گزارش وضعیت روشنایی", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "23-مشاهدات انجام شده در گشت زنی در راستای حفاظت از تاسیسات", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "24-احتمال بروز خطر یا حادثه", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "25-وضعیت سرقت درب", 
                    callback_data="electropump")], 
                    [InlineKeyboardButton(
                    "26-وضعیت اتاقک و حوضچه", 
                    callback_data="electropump")],
                    [InlineKeyboardButton(
                    "27-خطرات احتمالی سیل گیر بودن", 
                    callback_data="electropump")],
                    [InlineKeyboardButton(
                    "28-انجام اعمال غیر متعارف در محوطه تاسیسات (تخلیه زباله، رفت و آمد مشکوک، حفاری مشکوک و ...)", 
                    callback_data="electropump")],
                    [InlineKeyboardButton(
                    "29-انجام دستورالعمل های اعلام شده در حالت عادی و زمان بحران", 
                    callback_data="electropump")]
                    ]
# pass keyboard to reply_mark
        self.reply_mark_keyboard_choose_paymankar = InlineKeyboardMarkup(keyboard_choose_paymankar)
        self.reply_mark_keyboard_choose_time = InlineKeyboardMarkup(keyboard_choose_time)
        self.reply_mark_keyboard_name_bakhshha = InlineKeyboardMarkup(keyboard_name_bakhshha)
        self.reply_mark_keyboard_choose_office = InlineKeyboardMarkup(keyboard_choose_office)
        self.reply_mark_fill_online_mechanical = InlineKeyboardMarkup(keyboard_fill_online_mechanical)
        self.reply_mark_fill_online_electrical = InlineKeyboardMarkup(keyboard_fill_online_electrical)
        self.reply_mark_fill_online_administrator = InlineKeyboardMarkup(keyboard_fill_online_administrator)
        self.reply_mark_fill_online_quality_control = InlineKeyboardMarkup(keyboard_fill_online_quality_control)
        self.reply_mark_fill_online_debi_gage = InlineKeyboardMarkup(keyboard_fill_online_debi_gage)
        self.reply_mark_keyboard_name_bakhshha = InlineKeyboardMarkup(keyboard_name_bakhshha)
        # self.reply_mark_keyboard_name_well_jafarie = InlineKeyboardMarkup(keyboard_name_well_jafarie)
        self.reply_mark_keyboard_item_well = InlineKeyboardMarkup(keyboard_item_well)
        # self.reply_mark_keyboard_gozine_electropump = InlineKeyboardMarkup(keyboard_gozine_electropump)
        # self.reply_mark_keyboard_make_excel=InlineKeyboardMarkup(keyboard_make_excel)
        # self.reply_mark_keyboard_choose_month=InlineKeyboardMarkup(keyboard_choose_month)

# -----------------------------------define message_to_bot_function
    def messagetoUs(self, update: Update,  context: CallbackContext):
        if update.message.text=="1":
            bot.send_message(
                chat_id=update.message.chat_id,
                text=f'''{update.message.from_user.username}سلام 
                    به ربات ما خوش آمدید
                    این ربات جهت پر کردن آنلاین فرم های مورد نیاز کارفرما طراحی شده است
                    لطفا نام پیمانکاری که برنده ی قرارداد شدند را انتخاب کنید''', 
                reply_markup=self.reply_mark_keyboard_choose_paymankar
                )
        def choose_month(self):
            global month
            if update.message.reply_to_message.text=="اگر می خواهید از فایل های ارسالی ماه های قبل استفاده کنید نام ماه را ریپلای کنید و اگر فایل خام را می خواهید پر کنید عبارت خام را بنویسید":
                if update.message.text in dict_month.values():
                    month=update.message.text
                elif update.message.text=="خام":
                    now= jdatetime.datetime.now().strftime('%B')
                    month=dict_month.get(now)
            return (month)
        # print(choose_month(self))
        def choose_bakhsh(self):
            global bakhsh
            if  update.message.reply_to_message.text== "لطفا نام بخشی که برای آن کار می کنید را انتخاب کنید":
                if update.message.text == "ج":     
                    bakhsh="جعفریه"
                if update.message.text == "د":
                    bakhsh="دستجرد"      
                if update.message.text == "س":
                    bakhsh="سلفچگان"    
                if update.message.text == "ک":
                    bakhsh="کهک"
                if update.message.text == "م":
                    bakhsh="مرکزی"
                # else:
                    # bakhsh=bakhsh
            return(bakhsh)   
        def choose_office(self):
            global office
            if  update.message.reply_to_message.text== "لطفا نام دفتری که برای آن کار می کنید را انتخاب کنید":
                if update.message.text == "مکانیک":
                    office="واحد مکانیک"
                elif update.message.text == "برق":
                    office="واحد برق"
                elif update.message.text == "سرپرست":
                    office="واحد سرپرست"
                elif update.message.text == "کنترل کیفی":
                    office="واحد کنترل کیفی"
                elif update.message.text == "آب بدون درآمد":
                    office="واحد آب بدون درآمد"     
            return(office)
        def choose_local_location(self):
            global local_location
            if  update.message.reply_to_message.text== "لطفا فرم مورد نظر خود را انتخاب کرده و بر روی این پیام ریپلای کرده و نام روستا/شهر/محلی که اینفرم به آن تعلق دارد را بنویسید":
                # if update.message.text == "چاه":
                #     local_location="چاه"
                # elif update.message.text == "ایستگاه پمپاژ":
                #     local_location="ایستگاه پمپاژ"
                # elif update.message.text == "مخازن":
                #     local_location="مخازن"
                # elif update.message.text == "فضای سبز":
                #     local_location="فضای سبز"
                # elif update.message.text == "خط انتقال":
                #     local_location="خط انتقال"     
                # else:
                #     pass
                local_location=update.message.text
            return(local_location)
        if update.message.reply_to_message.text=='''،هر آیتمی را که انتخاب کردید به شکل عدد-انتخاب و توضیحات آن را در جلوی آن بنویسید.
                برای پرکردن فایل خام علامت * را وارد کنید(یک بار هم کافی است)''':
            if "1-" or"2-" or"3-" or"4-" or"5-" or"6-" or"7-" or"8-" or "9-" or "10-" or "11-" or "12-" or"13-" or"14-"or "15-" or"16-" or"17-" or"18-" or"19-"or"20-" or"21-" or"22-" or"23-" or"24-" or"25-" or"26-" or"27-" or"28-" in update.message.text:
                if "*" in update.message.text:
                    wb = load_workbook("36010101.xlsx")
                    ws = wb.active                
                    ws['b4']=(jdatetime.datetime.now()+jdatetime.timedelta(days=.147)).strftime("%Y-%m-%d , %H,%M'")
                else:
                    wb = load_workbook(choose_month(self)+" "+choose_bakhsh(self)+" "+choose_office(self)+" "+choose_local_location(self)+" "+".xlsx")
                    ws = wb.active                       
                    ws['b4']=(jdatetime.datetime.now()+jdatetime.timedelta(days=.147)).strftime("%Y-%m-%d , %H,%M'")
                if update.message.text.startswith("1-"):
                    ws['a8']=update.message.text.replace("1-",'')                    
                if update.message.text.startswith("2-"):
                    ws['b8']=update.message.text.replace("2-",'')
                if update.message.text.startswith("3-"):
                    ws['c8']=update.message.text.replace("3-",'')
                if update.message.text.startswith("4-"):
                    ws['d8']=update.message.text.replace("4-",'')
                if update.message.text.startswith("5-"):
                    ws['b12']=update.message.text.replace("5-",'')
                if update.message.text.startswith("6-"):
                    ws['b13']=update.message.text.replace("6-",'')
                if update.message.text.startswith("7-"):
                    ws['b14']=update.message.text.replace("7-",'')
                if update.message.text.startswith("8-"):
                    ws['b15']=update.message.text.replace("8-",'')
                if update.message.text.startswith("9-"):
                    ws['b16']=update.message.text.replace("9-",'')
                if update.message.text.startswith("10-"):
                    ws['b17']=update.message.text.replace("10-",'')
                if update.message.text.startswith("11-"):
                    ws['a21']=update.message.text.replace("11-",'')
                if update.message.text.startswith("12-"):
                    ws['b21']=update.message.text.replace("12-",'')
                if update.message.text.startswith("13-"):
                    ws['c21']=update.message.text.replace("13-",'')
                if update.message.text.startswith("14-"):
                    ws['d21']=update.message.text.replace("14-",'')
                if update.message.text.startswith("15-"):
                    ws['e21']=update.message.text.replace("15-",'')
                if update.message.text.startswith("16-"):
                    ws['a23']=update.message.text.replace("16-",'')
                if update.message.text.startswith("17-"):
                    ws['b23']=update.message.text.replace("17-",'')
                if update.message.text.startswith("18-"):
                    ws['c23']=update.message.text.replace("18-",'')
                if update.message.text.startswith("19-"):
                    ws['d23']=update.message.text.replace("19-",'')
                if update.message.text.startswith("20-"):
                    ws['e23']=update.message.text.replace("20-",'')
                if update.message.text.startswith("21-"):
                    ws['b26']=update.message.text.replace("21-",'')
                if update.message.text.startswith("22-"):
                    ws['b27']=update.message.text.replace("22-",'')
                if update.message.text.startswith("23-"):
                    ws['b28']=update.message.text.replace("23-",'')
                if update.message.text.startswith("24-"):
                    ws['b31']=update.message.text.replace("24-",'')
                if update.message.text.startswith("25-"):
                    ws['b32']=update.message.text.replace("25-",'')
                if update.message.text.startswith("26-"):
                    ws['a35']=update.message.text.replace("26-",'')
                if update.message.text.startswith("27-"):
                    ws['b35']=update.message.text.replace("27-",'')
                if update.message.text.startswith("28-"):
                    ws['c35']=update.message.text.replace("28-",'')
                if update.message.text.startswith("29-"):
                    ws['d35']=update.message.text.replace("29-",'')
                if update.message.text.startswith("30-"):
                    ws['a38']=update.message.text.replace("30-",'')
                wb.save(filename=
                # "/home/mahdimor01/"+name_bakhsh+"/"+name_office+"/"+name_checklist+"/"+
                (dict_month.get(jdatetime.datetime.now().strftime('%B'))+" "+choose_bakhsh(self)+" "+choose_office(self)+" "+choose_local_location(self)+" "+".xlsx"))
            else:
                pass
        print(choose_month(self))
        print(choose_bakhsh(self))
        print(choose_office(self))
        print(choose_local_location(self))
        print()
        
        # return("5")

# -----------------------------------define callbackquery_function-
    def query_btns(self, update: Update,  context: CallbackContext):
        query = update.callback_query
        # choose first query(فتحی)
        if query.data == "فتحی":
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                text="اگر می خواهید از فایل های ارسالی ماه های قبل استفاده کنید نام ماه را ریپلای کنید و اگر فایل خام را می خواهید پر کنید عبارت خام را بنویسید", 
                reply_markup=self.reply_mark_keyboard_choose_time
                )
        if query.data == "mahdavi":
            pass    
        if query.data in dict_month:
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                text="لطفا نام بخشی که برای آن کار می کنید را انتخاب کنید", 
                reply_markup=self.reply_mark_keyboard_name_bakhshha
                )
        elif query.data in list_name_bakhshha:
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                text="لطفا نام دفتری که برای آن کار می کنید را انتخاب کنید", 
                reply_markup=self.reply_mark_keyboard_choose_office
                )   
    
        if query.data == "fill_online_mechanical":
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                text="لطفا فرم مورد نظر خود را انتخاب کرده و بر روی این پیام ریپلای کرده و نام روستا/شهر/محلی که اینفرم به آن تعلق دارد را بنویسید", 
                reply_markup=self.reply_mark_fill_online_mechanical, 
                )
        # choose third query (want to fill_online_file for electrical)
        elif query.data == "fill_online_electrical":
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                reply_markup=self.reply_mark_fill_online_electrical, 
                text="لطفا فرم مورد نظر خود را انتخاب کرده و بر روی این پیام ریپلای کرده و نام روستا/شهر/محلی که اینفرم به آن تعلق دارد را بنویسید", 
                )
        # choose third query (want to fill_online_file for administrator)
        elif query.data == "fill_online_administrator":
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                text="لطفا فرم مورد نظر خود را انتخاب کرده و بر روی این پیام ریپلای کرده و نام روستا/شهر/محلی که اینفرم به آن تعلق دارد را بنویسید", 
                reply_markup=self.reply_mark_fill_online_administrator
                )
        elif query.data == "fill_online_quality_control":
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                text="لطفا فرم مورد نظر خود را انتخاب کرده و بر روی این پیام ریپلای کرده و نام روستا/شهر/محلی که اینفرم به آن تعلق دارد را بنویسید", 
                reply_markup=self.reply_mark_fill_online_quality_control
                )
        elif query.data == "fill_online_debi_gage":
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                text="لطفا فرم مورد نظر خود را انتخاب کرده و بر روی این پیام ریپلای کرده و نام روستا/شهر/محلی که اینفرم به آن تعلق دارد را بنویسید", 
                reply_markup=self.reply_mark_fill_online_debi_gage
                )
        if query.data=="36010101.xlsx":
            bot.send_message(
                chat_id=update.effective_message.chat_id, 
                text='''،هر آیتمی را که انتخاب کردید به شکل عدد-انتخاب و توضیحات آن را در جلوی آن بنویسید.
                برای پرکردن فایل خام علامت * را وارد کنید(یک بار هم کافی است)''',        
                reply_markup=self.reply_mark_keyboard_item_well
        )


# -----------------------------------------define innerfunction----
# ------------------------------------define main------------------
    def main(self):
        updater = Updater(token,use_context=True)
        updater.dispatcher.add_handler(MessageHandler(Filters.all, self.messagetoUs))
        updater.dispatcher.add_handler(CallbackQueryHandler(self.query_btns))
        updater.start_polling()
        updater.idle()
# -------------------------------write name_of_class_.main()-------
Mybot().main()